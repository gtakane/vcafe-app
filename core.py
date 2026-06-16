# -*- coding: utf-8 -*-
"""
core.py : データ読み込み・秘密情報・Discord送信などの中核ロジック。
Streamlit に依存しないので、単体でテスト・再利用できる。
"""
import os
import json
import re
from datetime import datetime, timezone, timedelta, date
from pathlib import Path

import pandas as pd


# ============================================================
# 秘密情報（Botトークン / Firebase鍵）の読み込み
#   優先順位: 環境変数 > secrets_local.json
#   ※ コードに直書きしない。secrets_local.json は .gitignore 済み。
# ============================================================
def load_secrets(base_dir: Path) -> dict:
    secrets = {}
    f = base_dir / "secrets_local.json"
    if f.exists():
        try:
            secrets = json.loads(f.read_text(encoding="utf-8"))
        except Exception:
            secrets = {}
    # 環境変数があれば上書き（環境変数を最優先）
    env_map = {
        "DISCORD_BOT_TOKEN": "discord_bot_token",
        "DISCORD_CHANNEL_ID": "discord_channel_id",
        "FIREBASE_KEY_PATH": "firebase_key_path",
    }
    for env_key, conf_key in env_map.items():
        v = os.environ.get(env_key)
        if v:
            secrets[conf_key] = v
    # FIREBASE_KEY_JSON 環境変数（Codespaces シークレット等）から JSON を直接読む
    firebase_json_str = os.environ.get("FIREBASE_KEY_JSON", "")
    if firebase_json_str:
        try:
            secrets["firebase_key_json"] = json.loads(firebase_json_str)
        except Exception:
            pass
    return secrets


def mask(value: str, keep: int = 4) -> str:
    """秘密情報を画面表示する時用のマスク。末尾数文字だけ残す。"""
    if not value:
        return "(未設定)"
    s = str(value)
    if len(s) <= keep:
        return "*" * len(s)
    return "*" * (len(s) - keep) + s[-keep:]


# ============================================================
# 勤怠レポート（kintai_report.py の出力 .xlsx）を読み込む
#   ※ 欠勤率/遅刻率は数式セルでキャッシュが無い場合があるため、
#     率は数値から自前で再計算する。
# ============================================================
def _to_num(series):
    return pd.to_numeric(series, errors="coerce").fillna(0)


def load_kintai_summary(path) -> pd.DataFrame:
    df = pd.read_excel(path, sheet_name="サマリー", engine="openpyxl")
    df = df[df["メイド名"].notna() & (df["メイド名"].astype(str).str.strip() != "")]
    for col in ["シフト数", "欠勤", "シフト変更", "追加シフト", "遅刻報告", "早退",
                "30分前ルール違反", "土日祝締切違反", "シフト提出遅れ(回)"]:
        if col in df.columns:
            df[col] = _to_num(df[col]).astype(int)
    if "平均連絡リードタイム(時間)" in df.columns:
        df["平均連絡リードタイム(時間)"] = pd.to_numeric(
            df["平均連絡リードタイム(時間)"], errors="coerce")
    # 率を再計算（シフト数=0 は NaN）
    df["欠勤率"] = (df["欠勤"] / df["シフト数"]).where(df["シフト数"] > 0)
    df["遅刻率"] = (df["遅刻報告"] / df["シフト数"]).where(df["シフト数"] > 0)
    return df.reset_index(drop=True)


def load_kintai_detail(path) -> pd.DataFrame:
    try:
        df = pd.read_excel(path, sheet_name="明細", engine="openpyxl")
    except Exception:
        return pd.DataFrame()
    return df[df["種別"].notna()].reset_index(drop=True) if "種別" in df.columns else df


def summarize_kpis(summary: pd.DataFrame, detail: pd.DataFrame) -> dict:
    k = {}
    k["メイド数"] = len(summary)
    k["総シフト数"] = int(summary["シフト数"].sum()) if "シフト数" in summary else 0
    k["総欠勤"] = int(summary["欠勤"].sum()) if "欠勤" in summary else 0
    k["平均欠勤率"] = (k["総欠勤"] / k["総シフト数"]) if k["総シフト数"] else 0.0
    if not detail.empty and "当日連絡" in detail.columns:
        k["当日連絡件数"] = int((detail["当日連絡"].astype(str).str.strip() != "").sum())
    else:
        k["当日連絡件数"] = 0
    if "30分前ルール違反" in summary.columns:
        k["30分前ルール違反"] = int(summary["30分前ルール違反"].sum())
    return k


# ============================================================
# シフト表（通常シフト）読み込み: 日付シート(YYYYMMDD)を取り込む
#   本番反映の前段。ここでは「読み取り＋プレビュー」までを担う。
# ============================================================
def list_shift_date_sheets(path) -> list:
    from openpyxl import load_workbook
    wb = load_workbook(path, read_only=True)
    return [s for s in wb.sheetnames if re.fullmatch(r"\d{8}", s)]


def load_shift_sheet(path, sheet_name) -> pd.DataFrame:
    # 数式セルがあるため data_only=True のキャッシュ値を読む。
    df = pd.read_excel(path, sheet_name=sheet_name, engine="openpyxl")
    cols = [c for c in ["roomId", "maidNickname", "maidId", "workshiftGroupId",
                        "openTime", "closeTime", "roomType", "roomModelId"]
            if c in df.columns]
    df = df[cols] if cols else df
    if "maidNickname" in df.columns:
        df = df[df["maidNickname"].notna() &
                (df["maidNickname"].astype(str).str.strip() != "")]
    return df.reset_index(drop=True)


# ============================================================
# Discord 送信（実行時にトークンが必要。ここでは関数だけ用意）
# ============================================================
def send_discord_message(token: str, channel_id: str, content: str) -> tuple:
    """チャンネルへメッセージ送信。成功で(True, 詳細)。"""
    import requests
    url = f"https://discord.com/api/v10/channels/{channel_id}/messages"
    headers = {"Authorization": f"Bot {token}"}
    r = requests.post(url, headers=headers, json={"content": content}, timeout=30)
    if r.status_code in (200, 201):
        return True, "送信成功"
    return False, f"失敗 status={r.status_code}: {r.text[:200]}"


# ============================================================
# Firestore 接続・データ取得・キャッシュ・集計
# ============================================================
JST = timezone(timedelta(hours=9))
FREE_TICKET_IDS  = frozenset({"trial10minutes"})
TICKET_PRICES = {
    "gokitaku30minutes": 840,
    "premiumGokitaku1":  960,
    "luckyGokitaku":     650,
}
RESERVATION_PRICE = 3920
CHEKI_PRICE       = 500
PAID_TICKET_IDS   = frozenset(TICKET_PRICES.keys())
_WEEKDAY_JP = ["月", "火", "水", "木", "金", "土", "日"]


def _calc_revenue(ticket_id: str, billed_coin: int, billed_reward: int,
                  room_type: str, is_test: bool) -> float:
    if is_test:
        return 0.0
    if room_type == "reservation":
        return float(RESERVATION_PRICE)
    if ticket_id in TICKET_PRICES:
        return float(TICKET_PRICES[ticket_id])
    return round((billed_coin + billed_reward) * 1.4, 0)


def _classify_type(ticket_id: str, is_test: bool) -> str:
    if is_test or ticket_id in FREE_TICKET_IDS:
        return "無料"
    return "有料"


def get_firestore_client(secrets: dict, base_dir: Path, env: str = "テスト"):
    try:
        import firebase_admin
        from firebase_admin import credentials, firestore as fb_fs
    except ImportError:
        raise ImportError(
            "firebase-admin が未インストールです。\n"
            "pip install firebase-admin を実行してください。"
        )
    app_name = f"vcafe_{env}"
    try:
        app = firebase_admin.get_app(app_name)
    except ValueError:
        key_json = secrets.get("firebase_key_json")
        if key_json and isinstance(key_json, dict):
            cred = credentials.Certificate(key_json)
        else:
            key_rel = secrets.get("firebase_key_path", "serviceAccountKey.json")
            key_path = (base_dir / key_rel).resolve()
            if not key_path.exists():
                raise FileNotFoundError(
                    f"サービスアカウント鍵が見つかりません: {key_path}\n"
                    "secrets_local.json の firebase_key_path を確認してください。"
                )
            cred = credentials.Certificate(str(key_path))
        app = firebase_admin.initialize_app(cred, name=app_name)
    return fb_fs.client(app)


def fetch_visits(db, start_jst: datetime, end_jst: datetime) -> pd.DataFrame:
    start_utc = start_jst.astimezone(timezone.utc)
    end_utc   = end_jst.astimezone(timezone.utc)
    q = (db.collection_group("userRecordVisits")
           .where("enterDateTime", ">=", start_utc)
           .where("enterDateTime", "<",  end_utc))

    rows = []
    for doc in q.stream():
        d = doc.to_dict()
        enter = d.get("enterDateTime")
        if enter is None:
            continue
        if getattr(enter, "tzinfo", None) is None:
            enter = enter.replace(tzinfo=timezone.utc)
        enter_jst     = enter.astimezone(JST)
        billed        = int(d.get("billedCoin") or 0)
        billed_reward = int(d.get("billedRewardPoint") or 0)
        ticket        = str(d.get("usedTicketItemId") or "ATCOIN")
        room_type     = str(d.get("roomType") or "")
        init_time     = int(d.get("initialTime") or 20)
        weight        = max(1, round(init_time / 20))
        rows.append({
            "doc_id":            doc.id,
            "userId":            doc.reference.parent.parent.id,
            "enterDateTime":     enter_jst.replace(tzinfo=None),
            "maidNickname":      str(d.get("maidNickname") or ""),
            "billedCoin":        billed,
            "billedRewardPoint": billed_reward,
            "usedTicketItemId":  ticket,
            "roomType":          room_type,
            "initialTime":       init_time,
            "visitWeight":       weight,
            "revenue":           0.0,
            "type":              "",
        })

    _COLS = ["doc_id", "userId", "enterDateTime", "maidNickname",
             "billedCoin", "billedRewardPoint", "usedTicketItemId", "roomType",
             "initialTime", "visitWeight", "revenue", "type"]
    if not rows:
        return pd.DataFrame(columns=_COLS)

    unique_uids = list({r["userId"] for r in rows if r["userId"]})
    test_user_ids: set = set()
    try:
        refs = [db.collection("users").document(uid) for uid in unique_uids]
        for udoc in db.get_all(refs):
            if udoc.exists and udoc.to_dict().get("testUser") is True:
                test_user_ids.add(udoc.id)
    except Exception:
        pass

    for r in rows:
        is_test = r["userId"] in test_user_ids
        r["revenue"] = _calc_revenue(
            r["usedTicketItemId"], r["billedCoin"], r["billedRewardPoint"],
            r["roomType"], is_test)
        r["type"] = _classify_type(r["usedTicketItemId"], is_test)

    df = pd.DataFrame(rows)
    df["enterDateTime"] = pd.to_datetime(df["enterDateTime"])
    return df.reset_index(drop=True)


def _ts_jst(ts):
    if ts is None:
        return None
    if hasattr(ts, "astimezone"):
        return ts.astimezone(JST).replace(tzinfo=None)
    return ts


def fetch_workshifts(db, start_jst: datetime, end_jst: datetime) -> pd.DataFrame:
    q = (db.collection_group("workshifts")
           .where("openTime", ">=", start_jst)
           .where("openTime", "<",  end_jst))
    rows = []
    for doc in q.stream():
        d = doc.to_dict()
        open_t  = _ts_jst(d.get("openTime"))
        close_t = _ts_jst(d.get("closeTime"))
        if open_t is None:
            continue
        serve_starts = [_ts_jst(t) for t in (d.get("serveStartTime") or []) if t is not None]
        serve_ends   = [_ts_jst(t) for t in (d.get("serveEndTime")   or []) if t is not None]
        rows.append({
            "shiftId":          doc.id,
            "roomId":           str(d.get("roomId") or ""),
            "maidNickname":     str(d.get("maidNickname") or ""),
            "maidId":           str(d.get("maidId") or ""),
            "workshiftGroupId": str(d.get("workshiftGroupId") or ""),
            "openTime":         open_t,
            "closeTime":        close_t,
            "serveStartTime":   serve_starts[0] if serve_starts else None,
            "serveEndTime":     serve_ends[-1]  if serve_ends   else None,
        })
    _COLS = ["shiftId", "roomId", "maidNickname", "maidId",
             "workshiftGroupId", "openTime", "closeTime",
             "serveStartTime", "serveEndTime"]
    if not rows:
        return pd.DataFrame(columns=_COLS)
    df = pd.DataFrame(rows)
    for c in ["openTime", "closeTime", "serveStartTime", "serveEndTime"]:
        df[c] = pd.to_datetime(df[c])
    return df.reset_index(drop=True)


def calc_shift_detail(df_shifts: pd.DataFrame) -> pd.DataFrame:
    _COLS = ["shiftId", "maidNickname", "openTime", "closeTime",
             "serveStartTime", "serveEndTime", "稼働時間", "遅刻分"]
    if df_shifts is None or df_shifts.empty:
        return pd.DataFrame(columns=_COLS)
    df = df_shifts.copy()
    df["actual_start"] = df["serveStartTime"].fillna(df["openTime"])
    df["actual_end"]   = df["serveEndTime"].fillna(df["closeTime"])
    df["稼働時間"] = ((df["actual_end"] - df["actual_start"])
                     .dt.total_seconds() / 3600).clip(lower=0)
    df["遅刻分"]  = ((df["actual_start"] - df["openTime"])
                     .dt.total_seconds().clip(lower=0) / 60)
    return df[_COLS].reset_index(drop=True)


def agg_shift_hours(df_detail: pd.DataFrame) -> pd.DataFrame:
    if df_detail is None or df_detail.empty:
        return pd.DataFrame(columns=["maidNickname", "稼働時間数", "遅刻合計分", "シフト数"])
    return (df_detail.groupby("maidNickname")
            .agg(稼働時間数=("稼働時間", "sum"),
                 遅刻合計分=("遅刻分",   "sum"),
                 シフト数=("shiftId",    "nunique"))
            .reset_index())


def fetch_cheki(db, start_jst: datetime, end_jst: datetime) -> pd.DataFrame:
    q = (db.collection_group("userAlbum")
           .where("date", ">=", start_jst)
           .where("date", "<",  end_jst))
    rows = []
    for doc in q.stream():
        d = doc.to_dict()
        dt = d.get("date")
        if dt is None:
            continue
        if hasattr(dt, "astimezone"):
            dt = dt.astimezone(JST).replace(tzinfo=None)
        rows.append({
            "doc_id":       doc.id,
            "userId":       doc.reference.parent.parent.id,
            "date":         dt,
            "maidNickname": str(d.get("maidNickname") or ""),
        })
    _COLS = ["doc_id", "userId", "date", "maidNickname"]
    if not rows:
        return pd.DataFrame(columns=_COLS)
    df = pd.DataFrame(rows)
    df["date"] = pd.to_datetime(df["date"])
    return df.reset_index(drop=True)


# ---- キャッシュ（parquet） ----

def _cache_path(base_dir: Path, name: str, env: str) -> Path:
    safe = "prod" if env == "本番" else "test"
    return base_dir / "data" / f"cache_{name}_{safe}.parquet"


def save_cache(df: pd.DataFrame, base_dir: Path, name: str, env: str) -> None:
    p = _cache_path(base_dir, name, env)
    p.parent.mkdir(exist_ok=True)
    df.to_parquet(p, index=False)


def load_cache(base_dir: Path, name: str, env: str):
    p = _cache_path(base_dir, name, env)
    if not p.exists():
        return None
    return pd.read_parquet(p)


def cache_mtime(base_dir: Path, name: str, env: str):
    p = _cache_path(base_dir, name, env)
    if not p.exists():
        return None
    return datetime.fromtimestamp(p.stat().st_mtime)


# ---- 集計 ----

def _biz_date(dt) -> date:
    """00:00〜01:59 は前日扱い（営業時間 19:00〜翌02:00）"""
    if dt.hour < 2:
        return (dt - timedelta(days=1)).date()
    return dt.date()


def agg_daily(df: pd.DataFrame) -> pd.DataFrame:
    d = df.copy()
    if "visitWeight" not in d.columns:
        d["visitWeight"] = 1
    d["date"]        = d["enterDateTime"].apply(_biz_date)
    d["paid_weight"] = d["visitWeight"] * (d["type"] == "有料").astype(int)
    g = d.groupby("date").agg(
        ご帰宅数=("visitWeight", "sum"),
        有料数=("paid_weight",  "sum"),
        売上=("revenue",       "sum"),
    ).reset_index()
    g["ご帰宅数"] = g["ご帰宅数"].astype(int)
    g["有料数"]   = g["有料数"].astype(int)
    g["平均単価"] = g.apply(
        lambda r: r["売上"] / r["有料数"] if r["有料数"] > 0 else 0.0, axis=1)
    return g


def agg_by_maid(df_visits: pd.DataFrame,
                df_cheki: pd.DataFrame = None,
                df_shift_hours: pd.DataFrame = None) -> pd.DataFrame:
    d = df_visits.copy()
    if "visitWeight" not in d.columns:
        d["visitWeight"] = 1
    d["paid_weight"]    = d["visitWeight"] * (d["type"] == "有料").astype(int)
    d["non_rsv_weight"] = d["visitWeight"] * (d["roomType"] != "reservation").astype(int)

    g = d.groupby("maidNickname").agg(
        ご帰宅数=("visitWeight",     "sum"),
        有料数=("paid_weight",      "sum"),
        予約数=("roomType",         lambda s: (s == "reservation").sum()),
        ご帰宅売上=("revenue",      "sum"),
        _non_rsv=("non_rsv_weight", "sum"),
    ).reset_index()
    g["ご帰宅数"] = g["ご帰宅数"].astype(int)
    g["有料数"]   = g["有料数"].astype(int)
    g["予約数"]   = g["予約数"].astype(int)
    g["平均単価"] = g.apply(
        lambda r: r["ご帰宅売上"] / r["有料数"] if r["有料数"] > 0 else 0.0, axis=1)

    if df_cheki is not None and not df_cheki.empty:
        cheki_cnt = df_cheki.groupby("maidNickname").size().reset_index(name="チェキ数")
        g = g.merge(cheki_cnt, on="maidNickname", how="left")
        g["チェキ数"] = g["チェキ数"].fillna(0).astype(int)
    else:
        g["チェキ数"] = 0
    g["チェキ売上"] = g["チェキ数"] * CHEKI_PRICE
    g["合計売上"]   = g["ご帰宅売上"] + g["チェキ売上"]

    if df_shift_hours is not None and not df_shift_hours.empty:
        sh = df_shift_hours[["maidNickname", "稼働時間数", "遅刻合計分"]].copy()
        g = g.merge(sh, on="maidNickname", how="left")
        g["稼働時間数"] = g["稼働時間数"].fillna(0).round(1)
        g["遅刻合計分"] = g["遅刻合計分"].fillna(0).round(0).astype(int)
        g["平均ご帰宅数"] = g.apply(
            lambda r: round(r["_non_rsv"] / r["稼働時間数"], 2)
            if r["稼働時間数"] > 0 else 0.0, axis=1)
    g = g.drop(columns=["_non_rsv"])
    return g.sort_values("ご帰宅数", ascending=False).reset_index(drop=True)


def agg_by_hour(df: pd.DataFrame) -> pd.DataFrame:
    d = df.copy()
    if "visitWeight" not in d.columns:
        d["visitWeight"] = 1
    d["hour"] = d["enterDateTime"].dt.hour
    g = d.groupby("hour").agg(ご帰宅数=("visitWeight", "sum")).reset_index()
    all_h = pd.DataFrame({"hour": range(24)})
    return all_h.merge(g, on="hour", how="left").fillna(0).astype({"ご帰宅数": int})


def agg_by_weekday(df: pd.DataFrame) -> pd.DataFrame:
    d = df.copy()
    if "visitWeight" not in d.columns:
        d["visitWeight"] = 1
    d["weekday_num"] = d["enterDateTime"].dt.weekday
    d["weekday"]     = d["weekday_num"].map(lambda x: _WEEKDAY_JP[x])
    return (d.groupby(["weekday_num", "weekday"])
             .agg(ご帰宅数=("visitWeight", "sum"))
             .reset_index()
             .sort_values("weekday_num")[["weekday", "ご帰宅数"]]
             .reset_index(drop=True))


def agg_by_ticket(df: pd.DataFrame) -> pd.DataFrame:
    d = df.copy()
    if "visitWeight" not in d.columns:
        d["visitWeight"] = 1
    return (d.groupby("usedTicketItemId")
             .agg(件数=("visitWeight", "sum"), 売上=("revenue", "sum"))
             .reset_index()
             .sort_values("件数", ascending=False)
             .astype({"件数": int})
             .reset_index(drop=True))


def agg_cheki(df_cheki: pd.DataFrame, df_visits: pd.DataFrame) -> dict:
    total  = len(df_cheki)
    visits = len(df_visits)
    by_maid = (df_cheki.groupby("maidNickname").size()
               .reset_index(name="チェキ数")
               .sort_values("チェキ数", ascending=False)
               .reset_index(drop=True))
    if not df_cheki.empty:
        tmp = df_cheki.copy()
        tmp["date"] = pd.to_datetime(tmp["date"]).dt.date
        daily = tmp.groupby("date").size().reset_index(name="チェキ数")
    else:
        daily = pd.DataFrame(columns=["date", "チェキ数"])
    return {
        "total":   total,
        "visits":  visits,
        "rate":    total / visits if visits > 0 else 0.0,
        "by_maid": by_maid,
        "daily":   daily,
    }


if __name__ == "__main__":
    # 簡易セルフテスト（アップロード済み実データで検証）
    up = Path("/mnt/user-data/uploads")
    s = load_kintai_summary(up / "勤怠レポート_2026-06.xlsx")
    d = load_kintai_detail(up / "勤怠レポート_2026-06.xlsx")
    print("サマリー行数:", len(s), "列:", list(s.columns))
    print(s[["メイド名", "シフト数", "欠勤", "欠勤率"]].head(5).to_string(index=False))
    print("KPI:", summarize_kpis(s, d))
    print("シフト日付シート:", list_shift_date_sheets(up / "シフト表.xlsx")[:5], "...")
    sh = load_shift_sheet(up / "シフト表.xlsx", "20260601")
    print("シフト(20260601) 行数:", len(sh), "列:", list(sh.columns))
